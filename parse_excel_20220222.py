#!/usr/bin/env python3
##coding=utf-8-sig
#上面那一行python 3 不需要
#Program:
#       從excel paser出資料, 依照需求, 產生客製報表
#History
#       2021/03/17 by 可愛的阿源 first release
#       2021/04/09 by 可愛的阿源 release
#       2021/05/12 by 可愛的阿源 release
#       2021/12/02 by 可愛的阿源 release
#       2022/02/18 by 可愛的阿源 release
#       2022/02/22 by 可愛的阿源 release
#Change log
#       2021/04/09 新增判斷上下班時間是否異常, 異常則填入顯示
#       2021/05/12 1. 修改由上班時間判定上班日期, 如果下班時間為04~12點, 判定為下班時間當日為上班
#                  2. 修改經由工作表名稱, 判斷新竹廠, 以及台南廠的上下班時間是否異常, 並且將產出的工作表, 命名為該廠
#                  3. 新增進場多次註解
#       2021/12/02 新增判斷self.check_work_time不能int時, 直接紀錄時間 (數值有1 day等字元)
#       2022/02/18 新增開啟csv時, 使用encoding="utf-8-sig" 解決亂碼問題
#       2022/02/22 新增paser excel時, 讀取到空白行會跳出迴圈

import csv  #windows python3.0 只需要import這個
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.comments import Comment

#使得sys.getdefaultencoding()的值為'utf-8', 下列兩行python3 不需要
#reload(sys)                       # reload才能調用setdefaultencoding方法 
#sys.setdefaultencoding('utf-8')   # 設置'utf-8'  

class ReadExcel:
    """用於讀寫excel資料"""

    def __init__(self, file_name, sheet_name1, sheet_name2, csv_change_file, map_csv_file, final_xlsx_file):
        """
        :param file_name: excle檔名
        :param sheet_name1: 表單名, 新竹
        :param sheet_name2: 表單名, 台南
        :param csv_change_file: excel轉csv檔,存檔名稱
        :param map_csv_file: 整理完姓名日期的資料, 產出csv的檔案
        :param final_xlsx_file: 最後整理完的資料, 產出xlsx的檔案
        """
        self.file_name = file_name
        self.sheet_name1 = sheet_name1
        self.sheet_name2 = sheet_name2
        self.csv_change_file = csv_change_file
        self.map_csv_file = map_csv_file
        self.final_xlsx_file = final_xlsx_file

    def open_excel(self):
        """開啟工作簿，選中表單"""
        self.wb = load_workbook(self.file_name)
        #判斷該檔案內擁有的工作表是那一個名稱, 用來決定使用哪一個場區判斷
        for sh_name in self.wb.sheetnames :
        #場區flag , 新竹廠1, 台南廠2
            if sh_name == self.sheet_name1 :
                self.sh = self.wb[self.sheet_name1]
                self.flag = 1
            elif sh_name == self.sheet_name2 :
                self.sh = self.wb[self.sheet_name2]
                self.flag = 2

    def close_excel(self):
        """關閉工作簿物件的方法"""
        self.wb.close()
    
    def paser_excel(self):
        """從excel取得資料的過程"""
        """讀取excel的資料, 將需要的資料parse出來, 存成csv檔案"""
        data_list = [] #創造一個空array
        self.name_set = set() # 設定名字集合, 去除重複姓名
        self.day_set = set() # 設定日期集合, 去除重複日期
        self.flag = 0 # 預設self.flag為0
        self.open_excel() # 開啟工作簿, 選中表單
        #判斷self.flag如果不為1或2, 表示指定的工作表不存在
        if not (self.flag == 1 or self.flag == 2) :
           print ("指定工作表名稱不存在")
           exit()
        #開啟輸出的CSV檔案
        with open(self.csv_change_file,"w",newline="",encoding="utf-8-sig") as csvfile:  #newline是3版本才有的
        #with open(self.csv_change_file,"wb") as csvfile:
            writer = csv.writer(csvfile) #建立CSV檔寫入器, python3 不需要delimiter
            #writer = csv.writer(csvfile,delimiter=',') #建立CSV檔寫入器
            for i in range(2, self.sh.max_row): #從第二行開始擷取資料
                # 取得姓名J
                self.user_name = self.sh.cell(row=i, column=10).value
                if str(self.user_name) == "None" : #2022/02/22 新增, 判斷是否為空白行
                #    print ('有空白')
                    continue
                # 取得上班打卡時間C
                self.work_start_time = self.sh.cell(row=i, column=3).value
                # 取得下班打卡時間D
                self.work_end_time = self.sh.cell(row=i, column=4).value
                # 計算總工作時數
                self.work_date = (self.work_end_time - self.work_start_time)
                # 取得上班日期, 取至月-天, 並且替換格式為月/日
                self.work_day_start = str(self.work_start_time).split()[0].split('-',1)[1].replace('-','/')
                # 取得下班日期, 取至月-天, 並且替換格式為月/日
                self.work_day_end = str(self.work_end_time).split()[0].split('-',1)[1].replace('-','/')
                # 取得總工作時數至秒 (H:M:s)
                self.work_time = str(self.work_date).split('.')[0]

                #---------以下為新增是否遲到早退判定------------------------------------------
                #工作日預設為上班打卡日期 (早班, 中班)
                self.work_day = self.work_day_start

                #取得上班時數判定是否有遲到早退(H)
                self.check_work_time = str(self.work_time).split(':')[0]

                # 取得上班打卡判別時間(H)
                self.check_work_start_time = str(self.work_start_time).split()[1].split(':')[0]

                # 取得下班打卡判別時間(H)
                self.check_work_end_time = str(self.work_end_time).split()[1].split(':')[0]

                try: #2021/12/02 新增判定self.check_work_time, 不能int時, 直接紀錄
                    #新竹場區判定
                    if self.flag == 1 :
                        #如果上班時數小於8,或者大於10, 上班打卡不在(06.15.22), 下班打卡不在(16.00.07) 判定為異常
                        if int(self.check_work_time) < 8 or int(self.check_work_time) >= 10 or self.check_work_start_time not in ('06','15','22') or self.check_work_end_time not in ('16','00','07'):
                            self.work_time = str(self.work_time) + " " + "上班時間:" + " " + str(self.work_start_time) + " " + "下班時間:" + " " + str(self.work_end_time)
                    #台南場區判定
                    elif self.flag == 2 :
                        #如果上班時數小於8,或者大於10, 上班打卡不在(06.15), 下班打卡不在(16.00) 判定為異常
                        if int(self.check_work_time) < 8 or int(self.check_work_time) >= 10 or self.check_work_start_time not in ('06','15') or self.check_work_end_time not in ('16','00'):
                            #上班打卡時間在23點, 下班打卡時間在07, 額外判斷
                            if self.check_work_start_time in ('23') and self.check_work_end_time in ('07') :
                                #分析上班打卡時間 (分鐘)
                                self.check_work_start_time_min = str(self.work_start_time).split()[1].split(':')[1]
                                #分析下班打卡時間 (分鐘)
                                self.check_work_end_time_min = str(self.work_end_time).split()[1].split(':')[1]
                                #判斷上班打卡時間如果在23點30分之前, 或者下班打卡時間在07點30分之後, 為正常上下班
                                if not (int(self.check_work_start_time_min) < 30 and int(self.check_work_end_time_min) >= 30) :
                                    self.work_time = str(self.work_time) + " " + "上班時間:" + " " + str(self.work_start_time) + " " + "下班時間:" + " " + str(self.work_end_time)
                            else :
                                self.work_time = str(self.work_time) + " " + "上班時間:" + " " + str(self.work_start_time) + " " + "下班時間:" + " " + str(self.work_end_time)
                except: #例外判斷
                    self.work_time = str(self.work_date) + " " + "上班時間:" + " " + str(self.work_start_time) + " " + "下班時間:" + " " + str(self.work_end_time)
                #-------------------------------------------------------------------------------

                # -------------判定晚班,工作日為下班打卡日期---------------------
                #判斷上班打卡日期, 與下班打卡日期不同時
                if self.work_day_start != self.work_day_end :
                    #如果下班打卡時間, 不是凌晨0點~5點區間, 上班打卡時間不是下午2點~4點區間, 則使用下班打卡的日期, 為工作日
                    if  self.check_work_end_time not in ('00','01','02','03','04','05') and self.check_work_start_time not in ('14','15','16'):
                        #設定工作日為下班打卡日期 (晚班)
                        self.work_day = self.work_day_end

                # ----------------------------------------------------------------
                
                #在CSV寫入一列資料
                writer.writerow([self.user_name,self.work_day,self.work_time])
    
                self.name_set.add(self.user_name) # 將取出的姓名加入集合,排除重複的
                self.day_set.add(self.work_day)   # 將取出的日期加入集合,排除重複的
                ####debug用
                #print ('姓名: ',self.user_name)
                #print ('上班打卡時間: ',self.work_start_time)
                #print ('下班打卡時間: ',self.work_end_time)
                #print ('總工作時數: ',self.work_date)
                #print ('上班打卡時間判別: ',self.check_work_start_time)
                #print ('取得上班時數: ',self.check_work_time)
                #print ('總工作時數: ',self.work_time)
                #print (" ")
            csvfile.close()
        self.name_list = list(self.name_set) # 將姓名轉為array
        self.day_list = list(self.day_set) # 將日期轉為array
        self.name_list.sort() # 將所有名子依照順序排列
        self.day_list.sort() # 將所有日期依照順序排列
        self.close_excel()
        ####debug用
        #print self.name_list
        #print self.day_list
        #print ("原始資料分析完畢 !!")  #windows python2.7顯示中文有問題
        print ("Original Excel Parse ok!")


    def make_map_csv_xlsx(self):

        """ 怕迷路的空間地圖
        姓名/日期 |  1/1  |  1/2  |  1/3  |
        ------------------------------
        葉某某    | 8小時 | 1小時 | 不爽來 |
        ------------------------------
        林某某    | 1小時 | 睡過頭| 不爽來 |

        row1/column1 | row1/column2 | row1/column3
        ------------------------------------------
        row2/column1 | row2,cloumn2 | row2/column3

        行 = row = 姓名 , 列 = column = 日期
        cell = 儲存格,   value = 值
        sheet = 工作表, workbook = 檔案
        """

        #開啟一個空的csv檔
        fp = open(self.map_csv_file, "w",encoding="utf-8-sig")
        #設定A1 cell的值
        fp.write("姓名/日期")
        #將所有日期, 依照順序寫入到同一行
        for day in self.day_list:
            fp.write(",")
            fp.write(day)
        #換行
        fp.write("\n")
        #將所有姓名, 依照順序寫入到同一列, 因此寫完就需要換行
        for name in self.name_list:
            fp.write(name)
            fp.write("\n")
        fp.close()

        #下列為將csv轉成xlsx
        wb = Workbook()
        ws = wb.active
        #依照旗標判定是那一個場區
        if self.flag == 1 :
            ws.title = "新竹廠"
        elif self.flag == 2 :
            ws.title = "台南廠"

        with open(self.map_csv_file, "r",encoding="utf-8-sig") as f:
            reader = csv.reader(f, delimiter=',')
            for row in reader:
                ws.append(row)
        #設定欄位大小
        #ws.row_dimensions.height = 15
        #ws.column_dimensions.width = 15
        ws.column_dimensions["A"].width = 15

        wb.save(self.final_xlsx_file)
        wb.close()

    def save_xlsx_from_mapcsv(self):
        """解析parse後的資料, 寫入xlsx檔"""

        #開啟xlsx表
        wb = load_workbook(self.final_xlsx_file)
        #依照旗標判定是那一個場區, 開啟該廠區名稱的工作表
        if self.flag == 1 :
            sh = wb['新竹廠']
        if self.flag == 2 :
            sh = wb['台南廠']
        #for row in sh.iter_rows():
        #    for cell in row:
        #        #cell.style.alignment.wrap_text=True
        #        cell.alignment = Alignment(wrapText=True)

        #開啟parse後儲存的檔案
        fp = open(self.csv_change_file, "r",encoding="utf-8-sig")
        #依行讀取內容
        for line in fp.readlines():
            #取得該行使用者名稱
            user_name = str(line).split(',')[0]
            #取得該行使用者日期
            work_day = str(line).split(',')[1]
            #取得該行使用者工作時數, 沒strip後面會有換行, 文字無法顯示!!!!
            work_time = str(line).split(',')[2].strip()
            #取得使用者名稱在陣列內的index, 用於xls內, 寫入定位
            user_index = self.name_list.index(user_name)
            #取得使用者工作日期在陣列內的index, 用於xls內, 寫入定位
            work_day_index = self.day_list.index(work_day)
            user_row = user_index + 2
            work_day_column = work_day_index + 2
          
            ####debug 用, 列出上列資訊
            #print user_name, user_index, user_row, work_day, work_day_index, work_day_column, work_time

            #----此區間為將資料寫入xlsx
            #判斷資料格的值是否為空
            check_data=sh.cell(user_row,work_day_column).value

            #如果資料格的值是空的
            if check_data == None :
                if len(work_time) >= 10 :
                    #取得使用者工作時數異常時資訊
                    split_work_time = str(work_time).split(' ',1)[0]
                    split_work_data1 = str(work_time).split(' ',1)[1].split(' ')[0]
                    split_work_data2 = str(work_time).split(' ',1)[1].split(' ')[1]
                    split_work_data3 = str(work_time).split(' ',1)[1].split(' ')[2]
                    split_work_data4 = str(work_time).split(' ',1)[1].split(' ')[3]
                    split_work_data5 = str(work_time).split(' ',1)[1].split(' ')[4]
                    split_work_data6 = str(work_time).split(' ',1)[1].split(' ')[5]
                    #將上述資料集合在一起
                    split_work_data = str(split_work_data1 + "\n" + split_work_data2 + "\n" + split_work_data3 + "\n" + split_work_data4 + "\n" + split_work_data5 + "\n" + split_work_data6)
                    #將使用者工作時數異常資訊, 變成註解插入
                    sh.cell(user_row,work_day_column).comment = Comment(text=split_work_data,author='')
                    #    #將儲存格背景變為紅色, 天藍色為1E90FF
                    #將儲存格背景變為紅色
                    fill = PatternFill("solid", fgColor="FF0000")
                    sh.cell(user_row,work_day_column).fill = fill
                    #將資訊寫入儲存格
                    sh.cell(user_row,work_day_column).value=split_work_time
                else:
                    #將資訊寫入儲存格
                    sh.cell(user_row,work_day_column).value=work_time
            #如果資料格的值不是空的
            else:
                #如果原來的資料格有註解, 將註解讀出, 如果沒有, 特別註明
                read_comment = sh.cell(user_row,work_day_column).comment
                #上一筆註解為空時
                if str(read_comment) == "None" :
                    read_comment = str("上一筆工作時段正常") + "\n" + "工作時數為:" + str(check_data) + "\n"
                #上一筆註解有資料時
                else:
                    read_comment = str(read_comment).replace('Comment:','').replace('by','').replace('進場多次','') + "\n"

                #將使用者工作時數異常資訊, 變成註解插入
                if len(work_time) >= 10 :
                  #取得使用者工作時數異常時資訊
                  split_work_time = str(work_time).split(' ',1)[0]
                  split_work_data1 = str(work_time).split(' ',1)[1].split(' ')[0]
                  split_work_data2 = str(work_time).split(' ',1)[1].split(' ')[1]
                  split_work_data3 = str(work_time).split(' ',1)[1].split(' ')[2]
                  split_work_data4 = str(work_time).split(' ',1)[1].split(' ')[3]
                  split_work_data5 = str(work_time).split(' ',1)[1].split(' ')[4]
                  split_work_data6 = str(work_time).split(' ',1)[1].split(' ')[5]
                  #將上述資料集合在一起
                  split_work_data = str(split_work_data1 + "\n" + split_work_data2 + "\n" + split_work_data3 + "\n" + split_work_data4 + "\n" + split_work_data5 + "\n" + split_work_data6)

                #將原有註解的資料, 加上此次異常的資訊
                comment_data = str(read_comment) + "\n" + str(split_work_data)
                #將資料寫入註解
                sh.cell(user_row,work_day_column).comment = Comment(text=comment_data,author='')
                #將儲存格背景變為橘色
                fill = PatternFill("solid", fgColor="FFA500")
                sh.cell(user_row,work_day_column).fill = fill
                #將資訊寫入儲存格
                sh.cell(user_row,work_day_column).value="進場多次"
                
    
        fp.close()

        #設定格式自動換行
        for row in sh.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrapText=True)

        #凍結第一行, 第一列
        sh.freeze_panes = 'B2'
        wb.save(self.final_xlsx_file)
        wb.close()
        #print ("資料彙整後的Excle, 已產生完畢 !!") #windows python2.7 中文顯示有問題
        print ("Creat new Excel file ok !!")

    #-------------------------------Debug測試用---------------------------
    def debug_csv(self):
        """將parse完的csv檔, 依照姓名排列, 並且刪除特殊字元, 簡易產出用"""
        #開啟輸出的CSV檔案
        with open("User_Parse.csv","w",newline="") as csvfile:  #newline是3版本才有的
        #with open('User_Parse.csv','wb') as csvfile:
            writer = csv.writer(csvfile) #建立CSV檔寫入器, python3 不需要delimiter
            #writer = csv.writer(csvfile,delimiter=',') #建立CSV檔寫入器
       # 依照姓名列表, 去搜尋名字
            for user in self.name_list:
                #開啟csv檔, 讀取資料並parse用戶, 將姓名相同的列出
                fp = open(self.csv_change_file, 'r')   
                for line in fp:
                    if line.count(user) > 0:
                        writer.writerow([line.strip()])
                        ####debug用, 這部份要輸出, 依照姓名排列的字串
                        #print (line.strip())
                fp.close()
        csvfile.close()
        print ("-------------Debug File---------------")
        #print ("依照姓名排列的CSV已完成(Debug用)") #windows python2.7 中文顯示有問題
        print ("The CSV list name sort file Creat!! (For Debug!)")

        #開啟一個空白檔案, 準備寫入處理完特殊字元的字串
        fp = open('Replaced.csv', "w")
        #開啟上面已經依照姓名排列的檔案
        fg = open('User_Parse.csv', 'r')
        #將已經依照姓名排列的檔案, 依序讀取出來
        for line in fg.readlines():
            #將" 字元替代成空白, 並將最左邊的空白刪除
            line2 = line.replace('"',' ').lstrip()
            #寫入檔案
            fp.write(line2)
            ####debug用, 這部份要輸出,姓名排列後, 刪除特殊字元的字串
            #print (line2)
        fp.close()
        fg.close()
        #print ("刪除特殊字元的CSV檔已完成(Debug用)") ##windows python2.7 中文顯示有問題
        print ("Del EOF ch csv file Creat !! (For Debug)")

if __name__ == '__main__':
    # 直接執行本檔案時執行，下面是一個應用例項
    # 需要讀取excel時直接呼叫ReadExcel類
    a = ReadExcel('123.xlsx','駐廠人員出入廠30天記錄查詢','EMS紀錄','Excel_Parse.csv','map.csv','final.xlsx')
    a.paser_excel()
    a.make_map_csv_xlsx()
    a.save_xlsx_from_mapcsv()
    #a.debug_csv()
